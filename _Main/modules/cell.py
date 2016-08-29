import openpyxl
import os

# Cellvalues = {}
wkbkpaths = {}
s = 0
active_wkbk_var=None


def open_wkbk(wkbk_path):
    try:
        global active_wkbk_var
    except:
        print "no active_wkbk_var (open_wkbk)"
    global wkbkpaths
    wkbk = openpyxl.load_workbook(wkbk_path)
    wkbkpaths[wkbk] = wkbk_path
    if active_wkbk not in globals():
        active_wkbk_var = wkbk
        return wkbk
    else:
        return wkbk


def open_wkbk_data_only(wkbk_path):
    try:
        global active_wkbk_var
    except:
        print "no active_wkbk_var (open_wkbk)"
    global wkbkpaths
    wkbk = openpyxl.load_workbook(wkbk_path, data_only=True)
    wkbkpaths[wkbk] = wkbk_path
    if active_wkbk not in globals():
        active_wkbk_var = wkbk
        return wkbk
    else:
        return wkbk


def active_wkbk(wkbk=None, data_only=False):
    l = 0
    global active_wkbk_var
    global active_sheet_var
    global s
    if wkbk is None:
        return active_wkbk_var
    elif not isinstance(wkbk, basestring):
        active_wkbk_var = wkbk
        sheetlist = active_wkbk_var.get_sheet_names()
        for sh in sheetlist:
            if "ironspread" in sh:
                sheetlist.remove(sh)
        active_sheet_var = active_wkbk_var.get_sheet_by_name(sheetlist[0])
        return wkbk
    try:
        global wkbkpaths
    except:
        print "No workbooks have been loaded (active_wkbk)"
        raise SystemExit(0)
    for wk in wkbkpaths:
        if wkbk in wkbkpaths[wk]:
            if data_only == False:
                active_wkbk_var = openpyxl.load_workbook(wkbkpaths[wk])
                sheetlist = active_wkbk_var.get_sheet_names()
                for sh in sheetlist:
                    if "ironspread" in sh:
                        sheetlist.remove(sh)
                active_sheet_var = active_wkbk_var.get_sheet_by_name(sheetlist[0])
                wkbkpaths[active_wkbk_var] = wkbkpaths[wk]
                l = 1
                break
            elif data_only == True:
                active_wkbk_var = openpyxl.load_workbook(wkbkpaths[wk], data_only=True)
                sheetlist = active_wkbk_var.get_sheet_names()
                for sh in sheetlist:
                    if "ironspread" in sh:
                        sheetlist.remove(sh)
                active_sheet_var = active_wkbk_var.get_sheet_by_name(sheetlist[0])
                wkbkpaths[active_wkbk_var] = wkbkpaths[wk]
                l = 1
                break
        else:
            try:
                if "new" in wkbkpaths[wkbk]:
                    active_wkbk_var = wkbk
                    sheetlist = active_wkbk_var.get_sheet_names()
                    for sh in sheetlist:
                        if "ironspread" in sh:
                            sheetlist.remove(sh)
                    active_sheet_var = active_wkbk_var.get_sheet_by_name(sheetlist[0])
                    wkbkpaths[active_wkbk_var] = "new%d.xlsx" % s
                    s += 1
                    l = 1
                    break
            except:
                pass
    if l == 0:
        print 'Workbook not found (active_wkbk)'
        raise SystemExit(0)


def active_sheet(sheet=None):
    if sheet is None:
        try:
            global active_sheet_var
            return active_sheet_var
        except:
            print "No active sheet (active_sheet)"
            raise SystemExit(0)
    try:
        global active_wkbk_var
    except:
        print "No active workbook (active_sheet)"
        raise SystemExit(0)
    active_sheet_var = active_wkbk_var.get_sheet_by_name(sheet)
    return active_sheet_var


class Cell(object):
    try:
        global active_sheet_var
    except:
        print "No active sheet (Cell object)"
        raise SystemExit(0)

    def __init__(self, x, y, sheet=None):
        if isinstance(x, basestring) or isinstance(x, unicode):
            self.sheet = x
            self.x = y
            self.y = sheet
        else:
            self.sheet = sheet
            self.x = x
            self.y = y

    @property
    def value(self):
        if self.sheet is not None:
            active_sheet(self.sheet)
        # Cellvalues[self.x, self.y] = active_sheet_var.cell(row=self.x, column=self.y).value
        return active_sheet_var.cell(row=self.x, column=self.y).value

    @value.setter
    def value(self, new_value):
        # Cellvalues[self.x,self.y]=new_value
        if self.sheet is not None:
            active_sheet(self.sheet)
        active_sheet_var.cell(row=self.x, column=self.y).value = new_value

    def is_empty(self):
        if active_sheet_var.cell(row=self.x, column=self.y).value is None:
            return True
        else:
            return False


def save(wkbk_path=None):
    global active_wkbk_var
    if wkbk_path == None:
        try:
            active_wkbk_var.save(wkbkpaths[active_wkbk_var])
            print "Saved at " + wkbkpaths[active_wkbk_var]
        except:
            print "There is no active workbook to save (save)"
            raise SystemExit(0)
    else:
        try:
            try:
                active_wkbk_var.save(wkbk_path)
                print "Saved at " + wkbk_path
            except:
                print "Could not save."
                # active_wkbk_var.save("C:/Users/Administrator/Desktop/Training Program Management/"+wkbk_path)
        except Exception as e:
            print "There is no active workbook to save or that workbook is open in the desktop (save)"
            print wkbk_path
            raise e


def all_sheets():
    global active_wkbk_var
    try:
        return active_wkbk_var.get_sheet_names()
    except:
        print "There is no active workbook to draw sheets from (all_sheets)"
        raise SystemExit(0)


def new_wkbk():
    global active_wkbk_var
    global s
    nb = openpyxl.Workbook()
    global wkbkpaths
    wkbkpaths[nb] = "new%d.xlsx" % s
    s += 1
    # print wkbkpaths
    result = active_wkbk(nb)
    active_wkbk_var = nb
    return nb


def close_wkbk(wkbk=None):
    try:
        global active_wkbk_var
    except:
        print "No active workbook(close_wkbk)"
        raise SystemExit(0)
    global active_wkbk_var
    if wkbk == None:
        save()
        pass
    else:
        if "Matrix" not in wkbk:
            save(wkbk)
    if wkbk == None:
        path = wkbkpaths[active_wkbk_var]
    else:
        for wk in wkbkpaths:
            if wkbk in wkbkpaths[wk]:
                path = wkbkpaths[wk]
                break
    to_delete = []
    for wk in wkbkpaths:
        if wkbkpaths[wk] == path:
            to_delete.append(wk)
    for wk in to_delete:
        del wkbkpaths[wk]
    try:
        if wkbk == None:
            del wkbkpaths[active_wkbk_var]
        else:
            del wkbkpaths[wkbk]
    except:
        pass
    if wkbk == active_wkbk_var or wkbk == None:
        delactive_wkbk_var()
        print active_wkbk_var


def delactive_wkbk_var():
    global active_wkbk_var
    global active_sheet_var
    active_wkbk_var = None
    active_sheet_var = None


def new_sheet(sheet_name=None):
    if sheet_name is None:
        try:
            global active_wkbk_var
            active_wkbk_var.create_sheet()
        except:
            print "No active workbook(new_sheet)"
            raise SystemExit(0)
    else:
        active_wkbk_var.create_sheet(title=sheet_name)


def autofit(sheet):
    # global active_sheet_var
    # try:
    #     column_widths = []
    #     for row in :
    #         for i, cell in enumerate(row):
    #             if len(column_widths) > i:
    #                 if len(cell) > column_widths[i]:
    #                     column_widths[i] = len(cell)
    #             else:
    #                 column_widths += [len(cell)]
    #
    #     for i, column_width in enumerate(column_widths):
    #         active_sheet_var.column_dimensions[get_column_letter(i+1)].width = column_width
    pass


    # mjrpath='C:/Users/Administrator/Desktop/test.xlsx'
    # mjr=open_wkbk(mjrpath)
    # active_sheet('Sheet1')
